"""
Service d'export Excel avec mise en forme
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class ExcelService:
    """Service d'export Excel avec mise en forme professionnelle"""
    
    # Couleurs
    COLOR_HEADER = "2C3E50"      # Bleu-gris foncé
    COLOR_TOTAL = "ECF0F1"       # Gris clair
    COLOR_SUCCESS = "27AE60"     # Vert
    
    @staticmethod
    def generate_factures_excel(documents, date_debut=None, date_fin=None, entreprise=None):
        """
        Génère un fichier Excel avec la liste des factures
        
        Args:
            documents: Liste de Documents (factures)
            date_debut: Date de début (optionnel)
            date_fin: Date de fin (optionnel)
            entreprise: Instance Entreprise
            
        Returns:
            Workbook: Fichier Excel
        """
        wb = Workbook()
        
        # Feuille 1 : Liste des factures
        ws_factures = wb.active
        ws_factures.title = "Factures"
        ExcelService._create_factures_sheet(ws_factures, documents, date_debut, date_fin, entreprise)
        
        # Feuille 2 : Statistiques
        ws_stats = wb.create_sheet("Statistiques")
        ExcelService._create_stats_sheet(ws_stats, documents, date_debut, date_fin, entreprise)
        
        return wb
    
    @staticmethod
    def _create_factures_sheet(ws, documents, date_debut, date_fin, entreprise):
        """Crée la feuille Liste des factures"""
        
        # En-tête entreprise
        if entreprise:
            ws['A1'] = entreprise.nom or "Mon Entreprise"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = "Liste des Factures"
            ws['A2'].font = Font(size=12, bold=True)
            
            if date_debut and date_fin:
                ws['A3'] = f"Période : du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
            
            start_row = 5
        else:
            start_row = 1
        
        # En-têtes colonnes
        headers = [
            "N° Facture",
            "Date",
            "Client",
            "Email",
            "Statut",
            "Total HT",
            "TVA",
            "Total TTC",
            "Date échéance"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=ExcelService.COLOR_HEADER, 
                                   end_color=ExcelService.COLOR_HEADER, 
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Filtrer par date
        filtered_docs = documents
        if date_debut:
            filtered_docs = [d for d in filtered_docs if d.date_emission >= date_debut]
        if date_fin:
            filtered_docs = [d for d in filtered_docs if d.date_emission <= date_fin]
        
        # Trier par date
        filtered_docs = sorted(filtered_docs, key=lambda d: d.date_emission, reverse=True)
        
        # Données
        row = start_row + 1
        total_ht = 0
        total_tva = 0
        total_ttc = 0
        
        for doc in filtered_docs:
            ws.cell(row=row, column=1, value=doc.numero)
            ws.cell(row=row, column=2, value=doc.date_emission.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=3, value=doc.client.nom_complet)
            ws.cell(row=row, column=4, value=doc.client.email or "")
            
            # Statut avec mapping
            statut_map = {
                'brouillon': 'Brouillon',
                'envoyee': 'Envoyée',
                'payee': 'Payée'
            }
            ws.cell(row=row, column=5, value=statut_map.get(doc.statut, doc.statut.capitalize()))
            
            # Montants
            ws.cell(row=row, column=6, value=float(doc.total_ht))
            ws.cell(row=row, column=7, value=float(doc.total_tva))
            ws.cell(row=row, column=8, value=float(doc.total_ttc))
            
            # Date échéance
            if doc.date_echeance:
                ws.cell(row=row, column=9, value=doc.date_echeance.strftime('%d/%m/%Y'))
            
            # Format montants
            for col in [6, 7, 8]:
                ws.cell(row=row, column=col).number_format = '#,##0.00 €'
            
            # Couleur statut
            if doc.statut == 'payee':
                ws.cell(row=row, column=5).font = Font(color=ExcelService.COLOR_SUCCESS, bold=True)
            
            total_ht += float(doc.total_ht)
            total_tva += float(doc.total_tva)
            total_ttc += float(doc.total_ttc)
            
            row += 1
        
        # Ligne de total
        row += 1
        ws.cell(row=row, column=5, value="TOTAL")
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=6, value=total_ht)
        ws.cell(row=row, column=7, value=total_tva)
        ws.cell(row=row, column=8, value=total_ttc)
        
        for col in [5, 6, 7, 8]:
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=ExcelService.COLOR_TOTAL,
                                   end_color=ExcelService.COLOR_TOTAL,
                                   fill_type="solid")
            cell.font = Font(bold=True)
            if col >= 6:
                cell.number_format = '#,##0.00 €'
        
        # Ajuster largeurs colonnes
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        
        # Bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_cells in ws.iter_rows(min_row=start_row, max_row=row, 
                                      min_col=1, max_col=len(headers)):
            for cell in row_cells:
                cell.border = thin_border
    
    @staticmethod
    def _create_stats_sheet(ws, documents, date_debut, date_fin, entreprise):
        """Crée la feuille Statistiques"""
        
        # Titre
        ws['A1'] = "STATISTIQUES"
        ws['A1'].font = Font(size=16, bold=True)
        
        if date_debut and date_fin:
            ws['A2'] = f"Période : du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        
        # Filtrer
        filtered_docs = documents
        if date_debut:
            filtered_docs = [d for d in filtered_docs if d.date_emission >= date_debut]
        if date_fin:
            filtered_docs = [d for d in filtered_docs if d.date_emission <= date_fin]
        
        # Calculs
        nb_total = len(filtered_docs)
        nb_brouillon = len([d for d in filtered_docs if d.statut == 'brouillon'])
        nb_envoyee = len([d for d in filtered_docs if d.statut == 'envoyee'])
        nb_payee = len([d for d in filtered_docs if d.statut == 'payee'])
        
        ca_total = sum(float(d.total_ttc) for d in filtered_docs)
        ca_paye = sum(float(d.total_ttc) for d in filtered_docs if d.statut == 'payee')
        ca_restant = sum(float(d.total_ttc) for d in filtered_docs if d.statut != 'payee')
        
        # Affichage
        row = 4
        
        # Nombre de factures
        ws[f'A{row}'] = "Nombre de factures"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        stats = [
            ("Total", nb_total),
            ("Brouillon", nb_brouillon),
            ("Envoyées", nb_envoyee),
            ("Payées", nb_payee)
        ]
        
        for label, value in stats:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            row += 1
        
        row += 1
        
        # Chiffre d'affaires
        ws[f'A{row}'] = "Chiffre d'affaires"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ca_stats = [
            ("CA Total", ca_total),
            ("CA Payé", ca_paye),
            ("CA Restant", ca_restant)
        ]
        
        for label, value in ca_stats:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            ws[f'C{row}'].number_format = '#,##0.00 €'
            if "Payé" in label:
                ws[f'C{row}'].font = Font(color=ExcelService.COLOR_SUCCESS, bold=True)
            row += 1
        
        # Ajuster largeurs
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
    
    @staticmethod
    def get_excel_filename(date_debut=None, date_fin=None):
        """Génère le nom du fichier Excel"""
        if date_debut and date_fin:
            return f"factures_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
        else:
            return f"factures_{datetime.now().strftime('%Y%m%d')}.xlsx"
